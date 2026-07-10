# Copyright Notice:
# Copyright 2016-2026 DMTF. All rights reserved.
# License: BSD 3-Clause License. For full text see link: https://github.com/DMTF/Redfish-Service-Validator/blob/main/LICENSE.md

import html as html_mod
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from redfish_service_validator import metadata
from redfish_service_validator.html_template import build_html_report
from redfish_service_validator.system_under_test import SystemUnderTest

_RSV_EXTRA_CSS = """
      /* Tally tables in sidebar */
      .tally-panel { margin-bottom: 0; }
      .tally-panel h3 { font-size: 10px; font-weight: 800; text-transform: uppercase; letter-spacing: 1px; color: #8a97aa; margin-bottom: 8px; }
      .tally-table { width: 100%; border-collapse: collapse; font-size: 11px; }
      .tally-table th { background: #f5f7fa; color: #2c3e50; font-weight: 700; padding: 5px 8px; text-align: left; border-bottom: 1px solid #dde3ec; }
      .tally-table td { padding: 4px 8px; border: none; border-bottom: 1px solid #f0f2f5; background: #fff; color: #444; }
      .tally-table tr:last-child td { border-bottom: none; }
      .tally-table tr:hover td { background: #f8fafc; }
      .tally-table td:last-child { text-align: right; font-weight: 700; color: #1a3a5c; width: 50px; }
      .tally-two-col { display: flex; flex-direction: column; gap: 14px; }

      /* Section heading */
      .section-heading { font-size: 11px; font-weight: 800; text-transform: uppercase; letter-spacing: 1px; color: #8a97aa; margin-bottom: 12px; padding-bottom: 6px; border-bottom: 1px solid #dde3ec; display: flex; align-items: center; gap: 8px; }
      .section-heading .sh-count { background: #e8f0fe; color: #0d6efd; font-size: 10px; padding: 1px 7px; border-radius: 10px; font-weight: 700; }

      /* Buttons */
      .btn { display: inline-flex; align-items: center; gap: 5px; padding: 5px 13px; border-radius: 5px; font-size: 11px; font-weight: 600; cursor: pointer; border: none; transition: background 0.15s, box-shadow 0.15s, transform 0.1s; user-select: none; white-space: nowrap; }
      .btn:active { transform: translateY(1px); }
      .btn-results { background: #0d6efd; color: #fff; box-shadow: 0 2px 5px rgba(13,110,253,0.3); }
      .btn-results:hover { background: #0b5ed7; }
      .btn-payload { background: #495057; color: #fff; box-shadow: 0 2px 5px rgba(73,80,87,0.3); }
      .btn-payload:hover { background: #343a40; }
      .btn-both { background: #198754; color: #fff; box-shadow: 0 2px 5px rgba(25,135,84,0.3); }
      .btn-both:hover { background: #157347; }
      .btn-group { display: flex; gap: 6px; flex-wrap: wrap; }

      /* Resource card */
      .resource-card { background: #fff; border: 1px solid #dde3ec; border-radius: 8px; margin-bottom: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.05); transition: box-shadow 0.15s; }
      .resource-card:hover { box-shadow: 0 3px 12px rgba(0,0,0,0.09); }
      .resource-card.hidden-by-filter { display: none; }
      .resource-header { display: flex; align-items: center; flex-wrap: wrap; gap: 10px; padding: 10px 16px; background: #f7f9fc; border-bottom: 1px solid #dde3ec; cursor: default; }
      .resource-uri { flex: 1; font-weight: 700; font-size: 13px; color: #0d1b2a; word-break: break-all; font-family: "Cascadia Code","Consolas",monospace; }
      .resource-type { font-size: 11px; color: #6c757d; font-style: italic; }
      .resource-badges { display: flex; gap: 5px; flex-wrap: wrap; align-items: center; }
      .badge { display: inline-block; padding: 2px 8px; border-radius: 20px; font-size: 11px; font-weight: 700; }
      .badge-status { background:#e8f0fe; color:#1565c0; }
      .badge-time { background:#f3e5f5; color:#6a1b9a; }

      /* Collapsible panels */
      .results { display: none; }
      .resultsShow { display: block; }

      /* Properties table */
      .prop-table { width: 100%; border-collapse: collapse; font-size: 12px; }
      .prop-table th { background: #f0f4f8; color: #2c3e50; font-weight: 700; padding: 7px 12px; text-align: left; border-bottom: 2px solid #dde3ec; }
      .prop-table td { padding: 5px 12px; border: none; border-bottom: 1px solid #f0f2f5; vertical-align: top; background: #fff; color: #333; word-break: break-word; }
      .prop-table tr:last-child td { border-bottom: none; }
      .prop-table tr:hover td { background: #f8fafc; }
      .prop-table td.res-pass { background:#eafaf1; color:#145a32; font-weight:700; text-align:center; width:90px; }
      .prop-table td.res-fail { background:#fdedec; color:#922b21; font-weight:700; text-align:center; width:90px; }
      .prop-table td.res-warn { background:#fef9e7; color:#7d6008; font-weight:700; text-align:center; width:90px; }
      .prop-table td.res-skip { background:#f5f7fa; color:#7f8c8d; font-style:italic; text-align:center; width:90px; }
      .msg-fail { color:#c0392b; font-size:11px; margin-top:2px; }
      .msg-warn { color:#b7770d; font-size:11px; margin-top:2px; }

      /* Panel toolbars */
      .panel-toolbar { display: flex; align-items: center; justify-content: flex-end; padding: 5px 10px; background: #eef2f7; border-bottom: 1px solid #dde3ec; gap: 6px; }
      .payload-toolbar { display: flex; align-items: center; justify-content: flex-end; padding: 5px 10px; background: #12151e; gap: 6px; }

      /* Payload */
      .payload-panel { background: #1e2130; color: #a8d8aa; font-family: "Cascadia Code","Consolas",monospace; font-size: 11px; padding: 14px 16px; overflow-x: auto; white-space: pre-wrap; word-wrap: break-word; max-height: 500px; overflow-y: auto; }

      /* Copy buttons */
      .btn-copy { display: inline-flex; align-items: center; gap: 4px; padding: 3px 10px; font-size: 11px; font-weight: 600; border-radius: 4px; cursor: pointer; border: 1px solid; transition: background 0.15s, color 0.15s; user-select: none; }
      .btn-copy-light { background:#fff; color:#555; border-color:#c8d3e0; }
      .btn-copy-light:hover { background:#e8f0fe; color:#0d6efd; border-color:#0d6efd; }
      .btn-copy-dark { background:transparent; color:#a8d8aa; border-color:#3a4a5c; }
      .btn-copy-dark:hover { background:#2a3a4c; color:#fff; }
      .btn-copy.copied { color:#27ae60!important; border-color:#27ae60!important; }

      /* No-match message */
      .filter-no-match { text-align:center; padding:40px; color:#aaa; font-size:14px; display:none; }
"""

_RSV_EXTRA_JS = r"""
      /* ── URI Filter ── */
      var allCards = [];
      window.addEventListener('DOMContentLoaded', function() {
        allCards = Array.from(document.querySelectorAll('.resource-card'));
        var tot = allCards.length;
        document.getElementById('filterCount').innerHTML = tot + ' / ' + tot;
        document.getElementById('totalCount').textContent = tot;
        document.getElementById('uriFilter').addEventListener('input', applyFilter);
      });
      function applyFilter() {
        var q = document.getElementById('uriFilter').value.trim().toLowerCase();
        document.getElementById('filterClear').style.display = q ? 'block' : 'none';
        var visible = 0, tot = allCards.length;
        allCards.forEach(function(c) {
          var uri = (c.querySelector('.resource-uri') || {}).textContent || '';
          var match = !q || uri.toLowerCase().indexOf(q) !== -1;
          c.classList.toggle('hidden-by-filter', !match);
          if (match) visible++;
        });
        document.getElementById('filterNoMatch').style.display =
          (visible === 0 && q) ? 'block' : 'none';
        document.getElementById('filterCount').innerHTML =
          '<b>' + visible + '</b> / ' + tot;
      }
      function clearFilter() {
        document.getElementById('uriFilter').value = '';
        applyFilter();
      }
      /* ── Copy helper ── */
      function copyToClipboard(btn, targetId) {
        var el = document.getElementById(targetId);
        var text = el ? el.innerText : '';
        _doCopy(btn, text);
      }
      function copyTableAsMarkdown(btn, tableContainerId) {
        var container = document.getElementById(tableContainerId);
        var table = container ? container.querySelector('table') : null;
        if (!table) { _doCopy(btn, ''); return; }
        var rows = Array.from(table.querySelectorAll('tr'));
        var mdLines = [];
        rows.forEach(function(tr, i) {
          var cells = Array.from(tr.querySelectorAll('th,td')).map(function(c) {
            return c.innerText.replace(/\n/g,' ').replace(/\|/g,'\\|').trim();
          });
          mdLines.push('| ' + cells.join(' | ') + ' |');
          if (i === 0) {
            var sep = cells.map(function(c, ci) {
              return ci === cells.length - 1 ? ':---:' : ':---';
            });
            mdLines.push('| ' + sep.join(' | ') + ' |');
          }
        });
        _doCopy(btn, mdLines.join('\n'));
      }
      function _doCopy(btn, text) {
        var orig = btn.innerHTML;
        function done() {
          btn.innerHTML = '&#10003; Copied!';
          btn.classList.add('copied');
          setTimeout(function() { btn.innerHTML = orig; btn.classList.remove('copied'); }, 1800);
        }
        if (navigator.clipboard) {
          navigator.clipboard.writeText(text).then(done).catch(function() { fallbackCopy(text); done(); });
        } else { fallbackCopy(text); done(); }
      }
      function fallbackCopy(text) {
        var ta = document.createElement('textarea');
        ta.value = text; ta.style.cssText = 'position:fixed;opacity:0';
        document.body.appendChild(ta); ta.select();
        document.execCommand('copy');
        document.body.removeChild(ta);
      }
"""

# ── Placeholder kept for any external code that may reference this name ──
html_template = None




def build_resource_header(
    uri, resource_type, uri_summary, payload_id, results_id, status_code=None, response_time=None
):
    """Builds the enterprise-styled resource card header row."""
    # Status code badge
    status_bg = "badge-status"
    if status_code is not None and status_code != 200:
        status_bg = "badge-fail"
    status_badge = '<span class="badge {}">{}</span>'.format(
        status_bg, "HTTP {}".format(status_code) if status_code is not None else "HTTP -"
    )
    # Response time badge
    time_badge = '<span class="badge badge-time">&#128336; {} ms</span>'.format(
        response_time if response_time is not None else "-"
    )
    return """
  <div class="resource-card">
    <div class="resource-header">
      <div>
        <div class="resource-uri">{uri}</div>
        <div class="resource-type">{rtype}</div>
      </div>
      <div class="resource-badges">{badges} {status} {rtime}</div>
      <div class="btn-group">
        <span class="btn btn-results"
          onclick="(function(btn){{var r=document.getElementById('{rid}'),p=document.getElementById('{pid}'),bp=btn.parentNode.querySelector('.btn-payload'),bb=btn.parentNode.querySelector('.btn-both');if(btn.getAttribute('data-state')==='on'){{r.classList.remove('resultsShow');btn.setAttribute('data-state','');btn.innerHTML='&#9776; Results';}}else{{r.classList.add('resultsShow');p.classList.remove('resultsShow');btn.setAttribute('data-state','on');btn.innerHTML='&#9650; Hide';if(bp){{bp.setAttribute('data-state','');bp.innerHTML='&#123;&#125; Payload';}}if(bb){{bb.setAttribute('data-state','');bb.innerHTML='&#9783; Both';}}}}}})(this)">
          &#9776; Results
        </span>
        <span class="btn btn-payload"
          onclick="(function(btn){{var r=document.getElementById('{rid}'),p=document.getElementById('{pid}'),br=btn.parentNode.querySelector('.btn-results'),bb=btn.parentNode.querySelector('.btn-both');if(btn.getAttribute('data-state')==='on'){{p.classList.remove('resultsShow');btn.setAttribute('data-state','');btn.innerHTML='&#123;&#125; Payload';}}else{{p.classList.add('resultsShow');r.classList.remove('resultsShow');btn.setAttribute('data-state','on');btn.innerHTML='&#9650; Hide';if(br){{br.setAttribute('data-state','');br.innerHTML='&#9776; Results';}}if(bb){{bb.setAttribute('data-state','');bb.innerHTML='&#9783; Both';}}}}}})(this)">
          &#123;&#125; Payload
        </span>
        <span class="btn btn-both"
          onclick="(function(btn){{var r=document.getElementById('{rid}'),p=document.getElementById('{pid}'),br=btn.parentNode.querySelector('.btn-results'),bp=btn.parentNode.querySelector('.btn-payload');if(btn.getAttribute('data-state')==='on'){{r.classList.remove('resultsShow');p.classList.remove('resultsShow');btn.setAttribute('data-state','');btn.innerHTML='&#9783; Both';}}else{{r.classList.add('resultsShow');p.classList.add('resultsShow');btn.setAttribute('data-state','on');btn.innerHTML='&#9650; Hide';if(br){{br.setAttribute('data-state','');br.innerHTML='&#9776; Results';}}if(bp){{bp.setAttribute('data-state','');bp.innerHTML='&#123;&#125; Payload';}}}}}})(this)">
          &#9783; Both
        </span>
      </div>
    </div>
""".format(
        uri=html_mod.escape(uri),
        rtype=html_mod.escape(resource_type),
        badges=uri_summary,
        status=status_badge,
        rtime=time_badge,
        pid=payload_id,
        rid=results_id,
    )


def build_resource_detail(results_id, results_str, payload_id, payload_str):
    """Builds the collapsible results + payload panels with copy buttons."""
    return """
    <div class="results" id="{rid}">
      <div class="panel-toolbar">
        <span class="btn-copy btn-copy-light" onclick="copyTableAsMarkdown(this, '{rid}-inner')" title="Copy as Markdown table">&#128203; Copy</span>
      </div>
      <div id="{rid}-inner">
      <table class="prop-table">
        <tr>
          <th style="width:30%">Property</th>
          <th>Value</th>
          <th style="width:90px">Result</th>
        </tr>
        {rows}
      </table>
      </div>
    </div>
    <div class="results" id="{pid}">
      <div class="payload-toolbar">
        <span class="btn-copy btn-copy-dark" onclick="copyToClipboard(this, '{pid}-inner')" title="Copy JSON payload">&#128203; Copy JSON</span>
      </div>
      <pre class="payload-panel" id="{pid}-inner">{payload}</pre>
    </div>
  </div>
""".format(rid=results_id, rows=results_str, pid=payload_id, payload=html_mod.escape(payload_str))


def build_not_tested_section(sut, uris):
    """Removed — Not Tested section no longer included in report."""
    return ""


def build_error_tally(error_classes, panel_title=None):
    """
    Creates a table of error/warning type counts.

    Args:
        error_classes: A dictionary containing the counts of types of errors
        panel_title: Optional heading for the tally panel

    Returns:
        The HTML string to insert in the results summary
    """
    if not error_classes:
        return ""
    rows = ""
    for etype in sorted(error_classes.keys()):
        rows += "<tr><td>{}</td><td>{}</td></tr>".format(html_mod.escape(etype + "s"), error_classes[etype])
    heading = "<h3>{}</h3>".format(html_mod.escape(panel_title)) if panel_title else ""
    return (
        '<div class="tally-panel">{}'
        '<table class="tally-table">'
        '<tr><th>Type</th><th style="text-align:right;width:80px">Count</th></tr>'
        "{}"
        "</table></div>"
    ).format(heading, rows)


def html_report(sut: SystemUnderTest, report_dir, time, tool_version, args=None):
    """
    Creates the HTML report for the system under test

    Args:
        sut: The system under test
        report_dir: The directory for the report
        time: The time the tests finished
        tool_version: The version of the tool
        args: The parsed CLI arguments dict

    Returns:
        The path to the HTML report
    """
    file = report_dir / datetime.strftime(time, "RedfishServiceValidatorReport_%m_%d_%Y_%H%M%S.html")
    html = ""

    # Build the error summary details — combined side-by-side panel
    error_tally = build_error_tally(sut._error_classes, "Failure Types")
    warning_tally = build_error_tally(sut._warning_classes, "Warning Types")
    if error_tally or warning_tally:
        combined_tally = '<div class="tally-two-col">{}{}</div>'.format(error_tally, warning_tally)
    else:
        combined_tally = ""

    # Build the URI results
    uris = sorted(list(sut._resources.keys()), key=str.lower)
    for index, uri in enumerate(uris):
        if not sut._resources[uri]["Validated"]:
            # Skip resources we didn't test
            # They might just be cached for reference link checks
            continue

        # Get the type info for the URI
        try:
            resource_type = sut._resources[uri]["Response"].dict["@odata.type"][1:].split(".")[0]
            try:
                resource_version = metadata.get_version(sut._resources[uri]["Response"].dict["@odata.type"])
                resource_version_str = "v{}.{}.{}".format(resource_version[0], resource_version[1], resource_version[2])
                resource_type += ", {}".format(resource_version_str)
            except:
                pass
        except:
            resource_type = "Unknown Resource Type"

        # Build the results summary for the URI
        uri_summary = '<span class="badge badge-pass">&#10003; Pass: {}</span>'.format(sut._resources[uri]["Pass"])
        if sut._resources[uri]["Warn"]:
            uri_summary += ' <span class="badge badge-warn">&#9888; Warn: {}</span>'.format(sut._resources[uri]["Warn"])
        if sut._resources[uri]["Fail"]:
            uri_summary += ' <span class="badge badge-fail">&#10007; Fail: {}</span>'.format(
                sut._resources[uri]["Fail"]
            )

        # Insert the URI results header
        results_id = "results{}".format(index)
        payload_id = "payload{}".format(index)
        html += build_resource_header(
            uri,
            resource_type,
            uri_summary,
            payload_id,
            results_id,
            status_code=sut._resources[uri].get("StatusCode"),
            response_time=sut._resources[uri].get("ResponseTime"),
        )

        # Insert the URI results details
        results_str = ""
        props = sorted(list(sut._resources[uri]["Results"]), key=str.lower)
        for prop in props:
            if prop == "":
                prop_str = "-"
            else:
                prop_str = prop[1:]
            result_class = ""
            raw_val = sut._resources[uri]["Results"][prop]["Value"]
            value_str = "<div>{}</div>".format(html_mod.escape(str(raw_val)) if raw_val is not None else "")
            if sut._resources[uri]["Results"][prop]["Result"] == "PASS":
                result_class = 'class="res-pass"'
            elif sut._resources[uri]["Results"][prop]["Result"] == "WARN":
                result_class = 'class="res-warn"'
                value_str += '<div class="msg-warn">{}</div>'.format(
                    html_mod.escape(sut._resources[uri]["Results"][prop]["Message"])
                )
            elif sut._resources[uri]["Results"][prop]["Result"] == "FAIL":
                result_class = 'class="res-fail"'
                value_str += '<div class="msg-fail">{}</div>'.format(
                    html_mod.escape(sut._resources[uri]["Results"][prop]["Message"])
                )
            elif sut._resources[uri]["Results"][prop]["Result"] == "SKIP":
                result_class = 'class="res-skip"'
            results_str += "<tr><td>{}</td><td>{}</td><td {}>{}</td></tr>".format(
                prop_str, value_str, result_class, sut._resources[uri]["Results"][prop]["Result"]
            )
        try:
            payload_str = json.dumps(
                sut._resources[uri]["Response"].dict, sort_keys=True, indent=4, separators=(",", ": ")
            )
        except:
            payload_str = "Malformed JSON"
        html += build_resource_detail(results_id, results_str, payload_id, payload_str)

    # Append the not-tested summary section (empty — removed per requirements)
    html += build_not_tested_section(sut, uris)

    # Build configuration rows for sidebar
    _config_keys = [
        "authtype",
        "certificatecheck",
        "config",
        "debugging",
        "ext_https_proxy",
        "logdir",
        "mockup",
        "payload",
        "requesttimeout",
        "serv_http_proxy",
        "token",
        "username",
        "verbose",
        "collectionlimit",
        "configuri",
        "ext_http_proxy",
        "forceauth",
        "metadatafilepath",
        "oemcheck",
        "requestattempts",
        "schema_directory",
        "serv_https_proxy",
        "uricheck",
        "usessl",
    ]
    config_rows_html = ""
    if args:
        for key in _config_keys:
            val = args.get(key, "")
            if val is None:
                val = ""
            elif isinstance(val, list):
                val = " ".join(str(v) for v in val)
            else:
                val = str(val)
            # Mask passwords/tokens
            if key in ("password", "token"):
                val = "********" if val else ""
            config_rows_html += "<tr><td>{}</td><td>{}</td></tr>".format(html_mod.escape(key), html_mod.escape(val))

    sidebar_extra = (
        '<div class="sidebar-section">{}</div>'.format(combined_tally)
        if combined_tally else ""
    )
    main_content = (
        '<div id="resourceList">{}</div>'
        '<div class="filter-no-match" id="filterNoMatch">No resources match your filter.</div>'
    ).format(html)
    main_prefix = (
        '<div class="section-heading">Resources Validated'
        '<span class="sh-count" id="totalCount"></span></div>'
    )

    with open(str(file), "w", encoding="utf-8") as fd:
        fd.write(
            build_html_report(
                page_title="Redfish Service Validator \u2014 Test Report",
                tool_title="Redfish Service Validator",
                filter_placeholder="Filter by URI\u2026",
                filter_count_label="resources",
                tool_link="https://github.com/DMTF/Redfish-Service-Validator",
                tool_repo="DMTF/Redfish-Service-Validator",
                tool_version=tool_version,
                generated_time=time.strftime("%c"),
                sut_host=html_mod.escape(str(sut.rhost)),
                sut_user=html_mod.escape(str(sut.username)),
                sut_password="********",
                sut_product=html_mod.escape(str(sut.product)),
                sut_manufacturer=html_mod.escape(str(sut.manufacturer)),
                sut_model=html_mod.escape(str(sut.model)),
                sut_firmware=html_mod.escape(str(sut.firmware_version)),
                pass_count=sut.pass_count,
                warn_count=sut.warn_count,
                fail_count=sut.fail_count,
                skip_count=sut.skip_count,
                sidebar_extra_html=sidebar_extra,
                config_rows_html=config_rows_html,
                extra_css=_RSV_EXTRA_CSS,
                main_prefix_html=main_prefix,
                main_content_html=main_content,
                extra_js=_RSV_EXTRA_JS,
            )
        )
    return file


def xlsx_report(sut: SystemUnderTest, report_dir, time, tool_version, args=None):
    """
    Creates an XLSX report for the system under test alongside the HTML report.

    Args:
        sut: The system under test
        report_dir: The directory for the report
        time: The time the tests finished
        tool_version: The version of the tool

    Returns:
        The path to the XLSX report
    """
    xlsx_file = report_dir / datetime.strftime(time, "RedfishServiceValidatorReport_%m_%d_%Y_%H%M%S.xlsx")

    # ── Colour palette ──────────────────────────────────────────────────
    C_HEADER_BG = "1565C0"
    C_HEADER_FG = "FFFFFF"
    C_PASS_BG = "D4EDDA"
    C_PASS_FG = "145A32"
    C_FAIL_BG = "F8D7DA"
    C_FAIL_FG = "7B241C"
    C_WARN_BG = "FFF3CD"
    C_WARN_FG = "7D6008"
    C_SKIP_BG = "F5F7FA"
    C_SKIP_FG = "7F8C8D"
    C_ALT_BG = "F0F4F8"
    C_URI_BG = "E8F0FE"
    C_SUMMARY_LBL = "2C3E50"

    def _fill(hex_color):
        return PatternFill(fill_type="solid", fgColor=hex_color)

    def _font(bold=False, color="000000", size=11):
        return Font(bold=bold, color=color, size=size)

    def _border():
        thin = Side(style="thin", color="000000")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _data():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    def _left():
        return Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── Header row helper ────────────────────────────────────────────────
    def _write_header(ws, row, cols):
        for col_idx, label in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.fill = _fill(C_HEADER_BG)
            cell.font = _font(bold=True, color=C_HEADER_FG, size=11)
            cell.alignment = _center()
            cell.border = _border()

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_view.showGridLines = False
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 45

    summary_rows = [
        ("Tool Version", str(tool_version)),
        ("Generated", time.strftime("%c")),
        ("Host", str(sut.rhost)),
        ("User", str(sut.username)),
        ("Product", str(sut.product)),
        ("Manufacturer", str(sut.manufacturer)),
        ("Model", str(sut.model)),
        ("Firmware", str(sut.firmware_version)),
        (None, None),
        ("Pass", sut.pass_count),
        ("Warning", sut.warn_count),
        ("Fail", sut.fail_count),
        ("Not Tested", sut.skip_count),
    ]

    result_colors = {
        "Pass": (C_PASS_BG, C_PASS_FG),
        "Warning": (C_WARN_BG, C_WARN_FG),
        "Fail": (C_FAIL_BG, C_FAIL_FG),
        "Not Tested": (C_SKIP_BG, C_SKIP_FG),
    }

    ws_summary.merge_cells("A1:B1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Redfish Service Validator — Summary"
    title_cell.fill = _fill(C_HEADER_BG)
    title_cell.font = _font(bold=True, color=C_HEADER_FG, size=13)
    title_cell.alignment = _center()
    title_cell.border = _border()
    ws_summary.row_dimensions[1].height = 22

    for r_idx, (label, value) in enumerate(summary_rows, start=2):
        if label is None:
            continue
        ca = ws_summary.cell(row=r_idx, column=1, value=label)
        cb = ws_summary.cell(row=r_idx, column=2, value=value)
        ca.font = _font(bold=True, color=C_SUMMARY_LBL)
        cb.font = _font()
        ca.alignment = _left()
        cb.alignment = _left()
        ca.border = _border()
        cb.border = _border()
        if label in result_colors:
            bg, fg = result_colors[label]
            ca.fill = _fill(bg)
            ca.font = _font(bold=True, color=fg)
            cb.fill = _fill(bg)
            cb.font = _font(bold=True, color=fg)
        else:
            ca.fill = _fill(C_ALT_BG)

    # Configuration section in Summary sheet
    if args:
        _config_keys = [
            "authtype",
            "certificatecheck",
            "config",
            "debugging",
            "ext_https_proxy",
            "logdir",
            "mockup",
            "payload",
            "requesttimeout",
            "serv_http_proxy",
            "token",
            "username",
            "verbose",
            "collectionlimit",
            "configuri",
            "ext_http_proxy",
            "forceauth",
            "metadatafilepath",
            "oemcheck",
            "requestattempts",
            "schema_directory",
            "serv_https_proxy",
            "uricheck",
            "usessl",
        ]
        # Blank separator row
        cfg_start = len([r for r in summary_rows if r[0] is not None]) + 3
        ws_summary.merge_cells("A{}:B{}".format(cfg_start, cfg_start))
        cfg_hdr = ws_summary["A{}".format(cfg_start)]
        cfg_hdr.value = "Configuration"
        cfg_hdr.fill = _fill(C_HEADER_BG)
        cfg_hdr.font = _font(bold=True, color=C_HEADER_FG, size=11)
        cfg_hdr.alignment = _center()
        cfg_hdr.border = _border()

        for ci, key in enumerate(_config_keys, start=cfg_start + 1):
            val = args.get(key, "")
            if val is None:
                val = ""
            elif isinstance(val, list):
                val = " ".join(str(v) for v in val)
            else:
                val = str(val)
            if key in ("password", "token"):
                val = "********" if val else ""
            ca = ws_summary.cell(row=ci, column=1, value=key)
            cb = ws_summary.cell(row=ci, column=2, value=val)
            ca.font = _font(bold=True, color=C_SUMMARY_LBL)
            cb.font = _font()
            ca.fill = _fill(C_ALT_BG)
            ca.alignment = _left()
            cb.alignment = _left()
            ca.border = _border()
            cb.border = _border()

    # ════════════════════════════════════════════════════════════════════
    # Sheet 2 — Detailed Results
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(title="Results")

    # Column widths: S#, URI, HTTP Status, Response Time (ms), Resource Type, Property, Value, Result, Message
    col_widths = [6, 55, 14, 18, 30, 35, 50, 12, 60]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.sheet_view.showGridLines = False
    _write_header(
        ws,
        1,
        ["S#", "URI", "HTTP Status", "Response Time (ms)", "Resource Type", "Property", "Value", "Result", "Message"],
    )
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    row_num = 2
    serial_num = 1
    uris = sorted(sut._resources.keys(), key=str.lower)

    result_fill = {
        "PASS": (_fill(C_PASS_BG), _font(bold=True, color=C_PASS_FG)),
        "FAIL": (_fill(C_FAIL_BG), _font(bold=True, color=C_FAIL_FG)),
        "WARN": (_fill(C_WARN_BG), _font(bold=True, color=C_WARN_FG)),
        "SKIP": (_fill(C_SKIP_BG), _font(color=C_SKIP_FG)),
    }

    for uri in uris:
        resource = sut._resources[uri]
        if not resource["Validated"]:
            continue

        # Resource type
        try:
            rtype = resource["Response"].dict["@odata.type"][1:].split(".")[0]
            try:
                rv = metadata.get_version(resource["Response"].dict["@odata.type"])
                rtype += " v{}.{}.{}".format(rv[0], rv[1], rv[2])
            except Exception:
                pass
        except Exception:
            rtype = "Unknown"

        props = sorted(resource["Results"].keys(), key=str.lower)
        first_row = True  # first property row of this URI — render bold

        for prop in props:
            prop_name = "-" if prop == "" else prop[1:]
            res_data = resource["Results"][prop]
            result = res_data.get("Result", "")
            raw_val = res_data.get("Value")
            message = res_data.get("Message", "") or ""

            val_str = str(raw_val) if raw_val is not None else ""
            # Strip "[Link to: ...]" wrapper — show the bare path instead
            if val_str.startswith("[Link to: ") and val_str.endswith("]"):
                val_str = val_str[len("[Link to: ") : -1]

            # Always fill URI and Resource Type on every row
            status_code = resource.get("StatusCode")
            resp_time = resource.get("ResponseTime")
            sn_cell = ws.cell(row=row_num, column=1, value=serial_num)
            uri_cell = ws.cell(row=row_num, column=2, value=uri)
            sc_cell = ws.cell(row=row_num, column=3, value=status_code if status_code is not None else "")
            rt_cell = ws.cell(row=row_num, column=4, value=resp_time if resp_time is not None else "")
            rtype_cell = ws.cell(row=row_num, column=5, value=rtype)
            prop_cell = ws.cell(row=row_num, column=6, value=prop_name)
            val_cell = ws.cell(row=row_num, column=7, value=val_str)
            res_cell = ws.cell(row=row_num, column=8, value=result)
            msg_cell = ws.cell(row=row_num, column=9, value=message)

            # S# cell
            sn_cell.fill = _fill("C7D9F1" if not first_row else C_URI_BG)
            sn_cell.font = _font(bold=first_row, color="0D1B2A")
            sn_cell.alignment = _data()
            sn_cell.border = _border()

            # URI / type columns — bold + darker background on first row of each URI group
            for c in (uri_cell, rtype_cell):
                c.fill = _fill("C7D9F1" if not first_row else C_URI_BG)
                c.font = _font(bold=first_row, color="0D1B2A")
                c.alignment = _data()
                c.border = _border()

            prop_cell.alignment = _data()
            prop_cell.border = _border()
            val_cell.alignment = _data()
            val_cell.border = _border()
            msg_cell.alignment = _data()
            msg_cell.border = _border()

            # HTTP Status cell — red if not 200
            sc_fill = _fill(C_FAIL_BG) if (status_code is not None and status_code != 200) else _fill("E8F0FE")
            sc_font = (
                _font(bold=True, color=C_FAIL_FG)
                if (status_code is not None and status_code != 200)
                else _font(bold=True, color="1565C0")
            )
            sc_cell.fill = sc_fill
            sc_cell.font = sc_font
            sc_cell.alignment = _data()
            sc_cell.border = _border()

            # Response Time cell — purple tint
            rt_cell.fill = _fill("F3E5F5")
            rt_cell.font = _font(color="6A1B9A")
            rt_cell.alignment = _data()
            rt_cell.border = _border()

            # Result cell colouring
            fill, font = result_fill.get(result, (_fill("FFFFFF"), _font()))
            res_cell.fill = fill
            res_cell.font = font
            res_cell.alignment = _data()
            res_cell.border = _border()

            first_row = False
            serial_num += 1
            row_num += 1

    # Auto-filter on header row
    ws.auto_filter.ref = "A1:{}1".format(get_column_letter(9))

    wb.save(str(xlsx_file))
    return xlsx_file
